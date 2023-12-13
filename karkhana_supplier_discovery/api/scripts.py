import frappe
import requests
import json


def supplier_to_supplier_master():
    success_migrations = []
    failed_migrations = []

    supplier_list = frappe.get_list("Supplier")
    # print(supplier_list)
    for i in supplier_list:
        supplier_doc = frappe.get_doc("Supplier",i.name)
        if not frappe.db.exists("Pre Stage", {"supplier_uuid": supplier_doc.unique_id}):
            try:
                create_pre_stage(supplier_doc)
                success_migrations.append({
                "supplier_name": supplier_doc.supplier_name,
                "status": "Success"
                })
            except Exception as e:

                # Handle the exception and add it to the failed migrations list
                reason = str(e)
                failed_migrations.append({
                "supplier_name": supplier_doc.supplier_name,
                "status": "Failed",
                "reason": reason
                })
    # save_migration_results_to_excel(success_migrations, failed_migrations,'/home/ubuntu/frappe-bench/apps/karkhana_supplier_discovery/migration_results.xlsx')
    return {
        "success":success_migrations,
        "failed":failed_migrations
    }


def create_pre_stage(supplier_doc):
    pre_stage_doc = frappe.new_doc("Pre Stage")
    pre_stage_doc.supplier_uuid = supplier_doc.unique_id
    pre_stage_doc.company_name = supplier_doc.supplier_name
    pre_stage_doc.email = supplier_doc.supplier_email
    pre_stage_doc.machines = supplier_doc.machine_capabilities
    pre_stage_doc.material_capabilities = supplier_doc.material_capabilities
    pre_stage_doc.company_gst = supplier_doc.gstin

    if supplier_doc.o_address or supplier_doc.o_city  or supplier_doc.o_state or supplier_doc.o_country or supplier_doc.o_pincode:
        pre_stage_doc.append("address",{
            "address_title":"Office Address",
            "address_line_1":supplier_doc.o_address,
            "city":supplier_doc.o_city,
            "state":supplier_doc.o_state,
            "country":supplier_doc.o_country,
            "pincode":supplier_doc.o_pincode

        })
    if supplier_doc.f_address or supplier_doc.f_city  or supplier_doc.f_state or supplier_doc.f_country or supplier_doc.f_pincode:
        pre_stage_doc.append("address",{
            "address_title":"Factory Address",
            "address_line_1":supplier_doc.f_address,
            "city":supplier_doc.f_city,
            "state":supplier_doc.f_state,
            "country":supplier_doc.f_country,
            "pincode":supplier_doc.f_pincode

        })
    pre_stage_doc.append("contact",{
            "contact_name":f"{supplier_doc.user_first_name} {supplier_doc.user_last_name}" if supplier_doc.user_first_name or supplier_doc.user_last_name else "" ,
            "email":supplier_doc.supplier_email,
            "phone":supplier_doc.supplier_mobile

        })
    certificate_list = frappe.get_list("Uploaded Certificates",filters={'supplier':supplier_doc.supplier_email})
    for i in certificate_list:
        certificate_doc = frappe.get_doc("Uploaded Certificates",i.name)
        pre_stage_doc.append("certificates",{
            "certificate_name":certificate_doc.certificate_name,
            "image":"/private"+str(certificate_doc.url)

        })
    pre_stage_doc.save()
    frappe.db.commit()

def save_migration_results_to_excel(success_migrations, failed_migrations, file_path):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    # Create a new Workbook
    wb = Workbook()

    # Create a new Excel sheet for successful migrations
    success_sheet = wb.create_sheet(title='Successful Migrations')
    
    # Create a new Excel sheet for failed migrations
    failed_sheet = wb.create_sheet(title='Failed Migrations')

    # Write headers to the Excel sheets for successful migrations
    if success_migrations:
        success_sheet.append(list(success_migrations[0].keys()))

    # Write headers to the Excel sheets for failed migrations
    if failed_migrations:
        failed_sheet.append(list(failed_migrations[0].keys()))

    # Write data to the Excel sheets
    for row in success_migrations:
        success_sheet.append(list(row.values()))

    for row in failed_migrations:
        failed_sheet.append(list(row.values()))

    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    # Save the Excel file
    wb.save(file_path)

def test():
    return 


